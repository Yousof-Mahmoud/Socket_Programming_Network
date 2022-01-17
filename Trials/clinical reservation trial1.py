import pandas as pd
import numpy as np

from openpyxl import load_workbook
msg = "LYRlCA"
wb = load_workbook(filename='clinic reservation.xlsx')
sheet = wb['clinic reservation']
value="DName"
DName = []

for i in range(1, sheet.max_row + 1):
    for j in range(sheet.max_column):
        if sheet[i][j].value:
            DName.append(sheet[i][j].value)
            if value in DName:
                print (value) 
            else:
                print (i,j)
